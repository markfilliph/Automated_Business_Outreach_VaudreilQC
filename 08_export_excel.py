"""
STEP 8: EXPORT EXCEL (v3: Hamilton Standard Format)

Creates a professionally formatted Excel workbook from the
Hamilton-standard CSV export (07_export.py output).

Two sheets:
  1. "Acquisition Leads" - main data, formatted and styled
  2. "Summary" - category breakdown, score distribution, pipeline stats

Input:  data/top_75_for_review.csv
Output: data/Vaudreuil_Acquisition_Leads.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))
from config import OUTPUT_FINAL, OUTPUT_EXCEL


# ── Styling constants ────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Arial", size=10)
SCORE_HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")    # Green
SCORE_MED_FILL = PatternFill("solid", fgColor="FFEB9C")     # Yellow
SCORE_LOW_FILL = PatternFill("solid", fgColor="FFC7CE")     # Red
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9")
)
SUMMARY_HEADER_FONT = Font(name="Arial", bold=True, size=12, color="1F4E79")
SUMMARY_LABEL_FONT = Font(name="Arial", bold=True, size=10)

# Column widths tuned for each field
COLUMN_WIDTHS = {
    "business_name": 30,
    "address": 28,
    "city": 18,
    "postal_code": 12,
    "website": 30,
    "phone": 18,
    "owner_name": 22,
    "owner_confidence": 14,
    "owner_source": 35,
    "industry": 18,
    "category_standardized": 25,
    "employee_range_estimate": 16,
    "revenue_range_estimate": 18,
    "sde_range_estimate": 18,
    "age_range_estimate": 16,
    "acquisition_fit_score": 14,
    "important_notes": 55,
}


def build_leads_sheet(wb, df):
    """Build the main leads data sheet."""
    ws = wb.active
    ws.title = "Acquisition Leads"

    # Freeze panes: header row + business_name column
    ws.freeze_panes = "B2"

    headers = list(df.columns)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            value = row[header]
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(header == "important_notes"))

            # Website as hyperlink
            if header == "website" and value and str(value).startswith("http"):
                cell.font = LINK_FONT
                cell.hyperlink = str(value)

            # Score conditional formatting
            if header == "acquisition_fit_score" and value:
                try:
                    score = int(float(value))
                    cell.alignment = Alignment(horizontal="center")
                    if score >= 70:
                        cell.fill = SCORE_HIGH_FILL
                    elif score >= 50:
                        cell.fill = SCORE_MED_FILL
                    else:
                        cell.fill = SCORE_LOW_FILL
                except (ValueError, TypeError):
                    pass

            # Owner confidence color coding
            if header == "owner_confidence":
                cell.alignment = Alignment(horizontal="center")
                if str(value).lower() == "high":
                    cell.fill = SCORE_HIGH_FILL
                elif str(value).lower() == "medium":
                    cell.fill = SCORE_MED_FILL
                elif str(value).lower() == "none":
                    cell.fill = SCORE_LOW_FILL

    # Set column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(header, 15)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"

    return ws


def build_summary_sheet(wb, df):
    """Build the summary statistics sheet."""
    ws = wb.create_sheet("Summary")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Vaudreuil-Dorion Acquisition Pipeline Summary")
    ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:D1")
    row += 1

    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=10, italic=True, color="808080")
    row += 2

    # Pipeline stats
    ws.cell(row=row, column=1, value="Pipeline Overview")
    ws.cell(row=row, column=1).font = SUMMARY_HEADER_FONT
    row += 1

    stats = [
        ("Total leads", len(df)),
        ("Score range", f"{df['acquisition_fit_score'].min()} to {df['acquisition_fit_score'].max()}"),
        ("Average score", f"{df['acquisition_fit_score'].mean():.0f}"),
        ("Owners identified", f"{len(df[df['owner_name'] != 'Not found'])}/{len(df)}"),
    ]
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = SUMMARY_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        row += 1

    row += 1

    # Score distribution
    ws.cell(row=row, column=1, value="Score Distribution")
    ws.cell(row=row, column=1).font = SUMMARY_HEADER_FONT
    row += 1

    scores = df["acquisition_fit_score"]
    dist = [
        ("Excellent (70+)", len(scores[scores >= 70]), SCORE_HIGH_FILL),
        ("Good (50-69)", len(scores[(scores >= 50) & (scores < 70)]), SCORE_MED_FILL),
        ("Fair (40-49)", len(scores[(scores >= 40) & (scores < 50)]), PatternFill("solid", fgColor="FFE0B2")),
        ("Below threshold (<40)", len(scores[scores < 40]), SCORE_LOW_FILL),
    ]
    for label, count, fill in dist:
        ws.cell(row=row, column=1, value=label).font = SUMMARY_LABEL_FONT
        c = ws.cell(row=row, column=2, value=count)
        c.font = DATA_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # Category breakdown
    ws.cell(row=row, column=1, value="Category Breakdown")
    ws.cell(row=row, column=1).font = SUMMARY_HEADER_FONT
    row += 1

    ws.cell(row=row, column=1, value="Category").font = SUMMARY_LABEL_FONT
    ws.cell(row=row, column=2, value="Count").font = SUMMARY_LABEL_FONT
    ws.cell(row=row, column=3, value="Avg Score").font = SUMMARY_LABEL_FONT
    row += 1

    categories = df.groupby("category_standardized").agg(
        count=("business_name", "count"),
        avg_score=("acquisition_fit_score", "mean")
    ).sort_values("count", ascending=False)

    for cat, data in categories.iterrows():
        ws.cell(row=row, column=1, value=cat).font = DATA_FONT
        ws.cell(row=row, column=2, value=data["count"]).font = DATA_FONT
        ws.cell(row=row, column=3, value=f"{data['avg_score']:.0f}").font = DATA_FONT
        row += 1

    row += 1

    # Industry breakdown
    ws.cell(row=row, column=1, value="Industry Breakdown")
    ws.cell(row=row, column=1).font = SUMMARY_HEADER_FONT
    row += 1

    industries = df["industry"].value_counts()
    for ind, count in industries.items():
        ws.cell(row=row, column=1, value=ind).font = DATA_FONT
        ws.cell(row=row, column=2, value=count).font = DATA_FONT
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    return ws


def main():
    print("=" * 60)
    print(" STEP 8: EXPORT EXCEL (Hamilton Standard Format)")
    print("=" * 60)

    df = pd.read_csv(OUTPUT_FINAL)
    print(f"  [INPUT] {len(df)} leads from {OUTPUT_FINAL}")

    wb = Workbook()

    # Sheet 1: Leads data
    print(f"  Building leads sheet...")
    build_leads_sheet(wb, df)

    # Sheet 2: Summary
    print(f"  Building summary sheet...")
    build_summary_sheet(wb, df)

    wb.save(OUTPUT_EXCEL)
    print(f"\n  [OUTPUT] Saved -> {OUTPUT_EXCEL}")
    print(f"  [OUTPUT] Sheets: Acquisition Leads, Summary")
    print(f"  [OUTPUT] {len(df)} leads, 17 fields, Hamilton standard format")


if __name__ == "__main__":
    main()
